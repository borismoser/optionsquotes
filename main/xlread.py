from openpyxl import load_workbook

workbook = load_workbook('sample.xlsx')  # opções: data_only = true, read_only = true
print(workbook.sheetnames)
sheet = workbook.active
print(sheet)
print(sheet.title)
print(sheet['a1'])
print(sheet['a1'].value)
print(sheet['f10'].value)
print(sheet.cell(row=10, column=6))  # f10 via index notation
print(sheet.cell(row=10, column=6).value)

print(sheet["A1:C2"])

# sheet["A"]    # Get all cells from column A
# sheet["A:B"]  # Get all cells for a range of columns
# sheet[5]      # Get all cells from row 5
# sheet[5:6]    # Get all cells for a range of rows

for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3):
    print(row)

for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3, values_only=True):
    print(row)

for column in sheet.iter_cols(min_row=1, max_row=2, min_col=1, max_col=3):
    print(column)

for column in sheet.iter_cols(min_row=1, max_row=2, min_col=1, max_col=3, values_only=True):
    print(column)

for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3):
    for cell in row:
        print(cell.value)

# atalho para .iter_rows sem argumentos (todas as linhas); há também .columns
# for row in sheet.rows:
#     print(row)


