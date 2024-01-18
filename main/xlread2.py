import json
from pprint import pprint
from openpyxl import load_workbook

workbook = load_workbook('sample.xlsx')  # opções: data_only = true, read_only = true
sheet = workbook.active

# for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
#     print(row)

# grab the columns with names: product_id, product_parent, product_title, product_category
# for value in sheet.iter_rows(min_row=2, min_col=4, max_col=7, values_only=True):
#     print(value)

products = {}
for row in sheet.iter_rows(min_row=2, min_col=4, max_col=7, values_only=True):
    product_id = row[0]
    product = {
        'parent': row[1],
        'title': row[2],
        'category': row[3]
    }
    products[product_id] = product

# Using json here to be able to format the output for displaying later
# pprint(json.dumps(products))
