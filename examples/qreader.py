from datetime import datetime

import openpyxl.utils.datetime
from openpyxl.reader.excel import load_workbook

# arquivo = "C:\\Users\\boris\\OneDrive\\Documentos\\Invest\\Opcoes_RTD.xlsx"
arquivo = "Opcoes_RTD.xlsx"
wb = load_workbook(filename=arquivo, data_only=True)
ws = wb.worksheets[1]
print(ws.title)

for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
    print(row)

venc = set()

for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
    for cell in row:
        venc.update(cell)
print(venc)
